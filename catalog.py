import os
import sys
import signal
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

def signal_handler(sig, frame):
    """Обработчик Ctrl+C для выхода из программы"""
    print('\n\nВыход из программы. До свидания!')
    sys.exit(0)

def get_directory_path():
    """Запрашивает путь к каталогу с проверкой существования"""
    while True:
        print("\n" + "="*60)
        print("Введите путь к каталогу (или 'exit'/'quit' для выхода)")
        print("Для отмены нажмите Ctrl+C")
        print("="*60)
        
        directory = input("Путь: ").strip().strip('"').strip("'")
        
        # Проверка на выход
        if directory.lower() in ['exit', 'quit', 'q', 'выход']:
            print("Выход из программы. До свидания!")
            sys.exit(0)
        
        # Проверка существования пути
        if not os.path.exists(directory):
            print(f"\n❌ ОШИБКА: Путь '{directory}' не существует!")
            print("Пожалуйста, проверьте путь и попробуйте снова.")
            continue
        
        # Проверка, что это директория
        if not os.path.isdir(directory):
            print(f"\n❌ ОШИБКА: '{directory}' не является каталогом!")
            print("Пожалуйста, укажите путь к каталогу.")
            continue
        
        # Проверка прав доступа
        if not os.access(directory, os.R_OK):
            print(f"\n❌ ОШИБКА: Нет прав на чтение каталога '{directory}'!")
            continue
        
        print(f"\n✅ Каталог '{directory}' принят.")
        return directory

def get_subdirectories(path):
    """Возвращает отсортированный список подкаталогов"""
    try:
        items = os.listdir(path)
        subdirs = []
        for item in items:
            full_path = os.path.join(path, item)
            if os.path.isdir(full_path):
                subdirs.append(item)
        return sorted(subdirs, key=str.lower)
    except PermissionError:
        return []
    except Exception as e:
        print(f"Ошибка при чтении каталога {path}: {e}")
        return []

def build_tree_structure(root_path):
    """
    Строит древовидную структуру каталогов.
    Каждая строка - это полный путь от корня до конечной папки.
    """
    rows = []
    
    def traverse(current_path, path_parts):
        """
        Рекурсивно обходит каталоги.
        path_parts - список названий папок от корня до текущей (включая текущую)
        """
        subdirs = get_subdirectories(current_path)
        
        if not subdirs:
            # Если нет подкаталогов - это конечная папка, добавляем строку
            rows.append(path_parts.copy())
            return
        
        # Для каждого подкаталога создаем новую ветку
        for subdir in subdirs:
            new_path_parts = path_parts + [subdir]
            subdir_path = os.path.join(current_path, subdir)
            traverse(subdir_path, new_path_parts)
    
    # Получаем подкаталоги первого уровня
    root_subdirs = get_subdirectories(root_path)
    
    if not root_subdirs:
        print("\n⚠️ В указанном каталоге нет подкаталогов.")
        return [["Нет подкаталогов"]]
    
    # Для каждого подкаталога первого уровня строим дерево
    for subdir in root_subdirs:
        subdir_path = os.path.join(root_path, subdir)
        traverse(subdir_path, [subdir])
    
    return rows

def save_to_excel(rows, root_path, output_file=None):
    """Сохраняет структуру в Excel файл"""
    if output_file is None:
        # Создаем имя файла на основе имени корневого каталога
        root_name = os.path.basename(root_path) or "root"
        # Убираем недопустимые символы из имени файла
        safe_name = "".join(c for c in root_name if c.isalnum() or c in (' ', '-', '_')).rstrip()
        if not safe_name:
            safe_name = "directory_structure"
        output_file = f"{safe_name}_structure.xlsx"
    
    # Создаем новую книгу Excel
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Структура каталогов"
    
    # Определяем максимальное количество столбцов (максимальная глубина)
    max_cols = max(len(row) for row in rows) if rows else 1
    
    # Стили для заголовков
    header_font = Font(name='Calibri', size=11, bold=True, color='FFFFFF')
    header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
    header_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    
    # Стили для данных
    data_alignment = Alignment(vertical='top', wrap_text=False)
    
    # Границы
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Добавляем заголовки столбцов
    # В первой ячейке пишем информацию о корневом каталоге
    ws.cell(row=1, column=1, value=f'Корневой каталог:\n{root_path}')
    ws.cell(row=1, column=1).font = header_font
    ws.cell(row=1, column=1).fill = header_fill
    ws.cell(row=1, column=1).alignment = header_alignment
    ws.cell(row=1, column=1).border = thin_border
    
    for col in range(2, max_cols + 1):
        cell = ws.cell(row=1, column=col, value=f'Уровень {col-1}')
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = thin_border
    
    # Заполняем данные
    for row_idx, row_data in enumerate(rows, 2):
        for col_idx, value in enumerate(row_data, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.alignment = data_alignment
            cell.border = thin_border
            
            # Чередуем цвета для четных/нечетных строк
            if row_idx % 2 == 0:
                cell.fill = PatternFill(start_color='D9E2F3', end_color='D9E2F3', fill_type='solid')
    
    # Настраиваем ширину столбцов
    for col in range(1, max_cols + 1):
        column_letter = openpyxl.utils.get_column_letter(col)
        # Находим максимальную длину в столбце
        max_length = 0
        for row in ws.iter_rows(min_col=col, max_col=col, values_only=True):
            for cell_value in row:
                if cell_value:
                    # Разделяем многострочные заголовки
                    for line in str(cell_value).split('\n'):
                        max_length = max(max_length, len(line))
        
        # Устанавливаем ширину (минимум 10, максимум 50)
        adjusted_width = min(max(max_length + 2, 10), 50)
        ws.column_dimensions[column_letter].width = adjusted_width
    
    # Замораживаем первую строку (заголовки)
    ws.freeze_panes = 'A2'
    
    # Добавляем автофильтр
    ws.auto_filter.ref = f'A1:{openpyxl.utils.get_column_letter(max_cols)}{len(rows)+1}'
    
    # Сохраняем файл
    try:
        wb.save(output_file)
        print(f"\n✅ Файл сохранен: {os.path.abspath(output_file)}")
        return output_file
    except PermissionError:
        alt_file = output_file.replace('.xlsx', '_new.xlsx')
        print(f"\n⚠️ Не удалось сохранить в '{output_file}' (возможно, файл открыт).")
        print(f"Сохраняем в '{alt_file}'")
        wb.save(alt_file)
        print(f"✅ Файл сохранен: {os.path.abspath(alt_file)}")
        return alt_file

def display_structure(rows, root_path):
    """Отображает структуру в консоли"""
    print("\n" + "="*60)
    print(f"СТРУКТУРА КАТАЛОГА: {root_path}")
    print("="*60)
    
    if not rows or rows == [["Нет подкаталогов"]]:
        print("Подкаталоги отсутствуют.")
        return
    
    max_cols = max(len(row) for row in rows)
    
    # Выводим первые 30 строк для примера
    print(f"\nПервые 30 строк структуры:")
    print("-" * 60)
    
    for i, row in enumerate(rows[:30]):
        # Показываем путь с отступами
        path_str = " > ".join(str(x) for x in row if x)
        print(f"{i+1:3d}. {path_str}")
    
    if len(rows) > 30:
        print(f"\n... и еще {len(rows) - 30} строк (полная структура в Excel)")

def main():
    # Устанавливаем обработчик Ctrl+C
    signal.signal(signal.SIGINT, signal_handler)
    
    print("\n" + "="*60)
    print("ПРОГРАММА СОХРАНЕНИЯ СТРУКТУРЫ КАТАЛОГОВ В EXCEL")
    print("="*60)
    print("\nЭта программа создает Excel-файл с иерархией подкаталогов.")
    print("Каждая строка - это полный путь от корня до конечной папки:")
    print("  Столбец A: подкаталоги 1-го уровня")
    print("  Столбец B: подкаталоги 2-го уровня")
    print("  Столбец C: подкаталоги 3-го уровня и т.д.")
    print("\n📌 Дублирования названий нет - каждая папка на своем уровне.")
    print("\nДля выхода введите 'exit', 'quit' или нажмите Ctrl+C")
    
    # Запрашиваем путь к каталогу
    root_path = get_directory_path()
    
    print("\n⏳ Анализирую структуру каталогов...")
    
    # Строим структуру
    rows = build_tree_structure(root_path)
    
    # Отображаем предварительный результат
    display_structure(rows, root_path)
    
    # Сохраняем в Excel
    output_file = save_to_excel(rows, root_path)
    
    print("\n" + "="*60)
    print(f"📊 Статистика:")
    print(f"   • Количество конечных папок (строк): {len(rows)}")
    if rows and rows != [["Нет подкаталогов"]]:
        print(f"   • Максимальная глубина (уровней): {max(len(row) for row in rows)}")
    print("="*60)
    
    # Предлагаем обработать еще один каталог
    while True:
        print("\n" + "-"*40)
        again = input("Обработать другой каталог? (да/нет): ").strip().lower()
        if again in ['да', 'yes', 'y', 'д']:
            main()
            break
        elif again in ['нет', 'no', 'n', 'н']:
            print("\nСпасибо за использование программы! До свидания!")
            break
        else:
            print("Пожалуйста, ответьте 'да' или 'нет'.")

if __name__ == "__main__":
    try:
        main()
    except KeyboardInterrupt:
        print('\n\nПрограмма прервана пользователем. До свидания!')
        sys.exit(0)