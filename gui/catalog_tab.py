import os
import queue
import subprocess
import threading
import tkinter as tk
from tkinter import ttk
from typing import Callable
import customtkinter as ctk

from core.scanner import FileInfo, MediaDirInfo
from core.classifier import Classifier
from utils import cache as cache_utils
from ai.taxonomy import get_active as _get_taxonomy


class CatalogTab(ctk.CTkFrame):
    def __init__(self, parent, get_files: Callable[[], list[FileInfo | MediaDirInfo]],
                 get_scan_root: Callable[[], str] | None = None):
        super().__init__(parent, fg_color="transparent")
        self._get_files = get_files
        self._get_scan_root = get_scan_root
        self._entries: list[dict] = []
        self._iid_to_idx: dict[str, int] = {}
        self._idx_to_iid: dict[int, str] = {}
        self._cache: dict = {}
        self._classifier: Classifier | None = None
        self._queue: queue.Queue = queue.Queue()
        self._edit_idx: int = -1
        self._classify_gen: int = 0
        self._meta_only_var = tk.BooleanVar(value=False)
        self._preselecting = False
        self._taxonomy = self._build_taxonomy(_get_taxonomy()[0])

        self._build_ui()
        self._poll_queue()

    # ------------------------------------------------------------------ build

    def _build_ui(self):
        # --- row 1: load buttons ---
        top = ctk.CTkFrame(self, fg_color="transparent")
        top.pack(fill="x", padx=12, pady=(10, 2))

        self._load_btn = ctk.CTkButton(
            top, text="Загрузить из сканирования", width=215,
            command=self._load_from_scan)
        self._load_btn.pack(side="left", padx=(0, 8))

        self._load_disk_btn = ctk.CTkButton(
            top, text="Загрузить с диска", width=150,
            fg_color="gray30", hover_color="gray25",
            command=self._load_from_disk)
        self._load_disk_btn.pack(side="left", padx=(0, 8))

        self._update_btn = ctk.CTkButton(
            top, text="Обновить каталог", width=160, state="disabled",
            fg_color="#2a5a2a", hover_color="#1e421e",
            command=self._start_incremental_update)
        self._update_btn.pack(side="left", padx=(0, 8))

        self._export_btn = ctk.CTkButton(
            top, text="Экспорт в Excel", width=150, state="disabled",
            fg_color="gray30", hover_color="gray25",
            command=self._export_excel)
        self._export_btn.pack(side="right")

        self._stats_label = ctk.CTkLabel(top, text="Файлы не загружены",
                                          text_color="gray60")
        self._stats_label.pack(side="right", padx=12)

        # --- row 2: classify buttons ---
        top2 = ctk.CTkFrame(self, fg_color="transparent")
        top2.pack(fill="x", padx=12, pady=(0, 4))

        self._classify_btn = ctk.CTkButton(
            top2, text="Классифицировать", width=160, state="disabled",
            command=self._start_classify)
        self._classify_btn.pack(side="left", padx=(0, 8))

        self._stop_btn = ctk.CTkButton(
            top2, text="Стоп", width=80, state="disabled",
            fg_color="#555", hover_color="#444",
            command=self._stop_classify)
        self._stop_btn.pack(side="left", padx=(0, 16))

        self._meta_only_cb = ctk.CTkCheckBox(
            top2, text="Только метаданные (быстро)",
            variable=self._meta_only_var,
            command=self._update_classify_button)
        self._meta_only_cb.pack(side="left", padx=(0, 16))

        self._reset_btn = ctk.CTkButton(
            top2, text="Сбросить классификацию", width=200, state="disabled",
            fg_color="#6b1a1a", hover_color="#551515",
            command=self._reset_classifications)
        self._reset_btn.pack(side="left")

        # --- filter + search ---
        flt = ctk.CTkFrame(self, fg_color="transparent")
        flt.pack(fill="x", padx=12, pady=(0, 4))

        ctk.CTkLabel(flt, text="Категория:", anchor="w").pack(side="left")
        self._filter_var = tk.StringVar(value="Все")
        self._filter_box = ctk.CTkComboBox(
            flt, variable=self._filter_var, values=["Все"],
            width=220, command=lambda _: self._apply_filter())
        self._filter_box.pack(side="left", padx=(4, 16))

        ctk.CTkLabel(flt, text="Поиск:", anchor="w").pack(side="left")
        self._search_var = tk.StringVar()
        self._search_var.trace_add("write", lambda *_: self._apply_filter())
        ctk.CTkEntry(flt, textvariable=self._search_var, width=200,
                     placeholder_text="по названию, автору, файлу...").pack(side="left", padx=4)

        # --- progress ---
        prog = ctk.CTkFrame(self, fg_color="transparent")
        prog.pack(fill="x", padx=12, pady=(0, 4))

        self._progress = ctk.CTkProgressBar(prog)
        self._progress.pack(fill="x")
        self._progress.set(0)

        self._status_label = ctk.CTkLabel(
            prog, text="", text_color="gray60",
            anchor="w", font=ctk.CTkFont(size=11))
        self._status_label.pack(fill="x", pady=(2, 0))

        # --- main area: table (left) + right edit panel ---
        main_area = ctk.CTkFrame(self, fg_color="transparent")
        main_area.pack(fill="both", expand=True, padx=12, pady=(0, 10))

        table_frame = ctk.CTkFrame(main_area)
        table_frame.pack(side="left", fill="both", expand=True)
        self._tree = self._create_tree(table_frame)

        right = ctk.CTkFrame(main_area, width=355)
        right.pack(side="right", fill="y", padx=(8, 0))
        right.pack_propagate(False)
        self._build_right_panel(right)

    def _create_tree(self, parent) -> ttk.Treeview:
        style = ttk.Style()
        style.theme_use("default")
        bg, fg, sel, head = "#2b2b2b", "#e0e0e0", "#1a5276", "#1f538d"

        style.configure("CAT.Treeview",
                        background=bg, foreground=fg, fieldbackground=bg,
                        rowheight=22, borderwidth=0, font=("Segoe UI", 10))
        style.configure("CAT.Treeview.Heading",
                        background=head, foreground="white",
                        font=("Segoe UI", 10, "bold"), relief="flat")
        style.map("CAT.Treeview", background=[("selected", sel)])
        style.map("CAT.Treeview.Heading", background=[("active", "#144578")])

        cols = ("title", "author", "year", "publisher", "category",
                "name", "ext", "size_mb", "status")
        tree = ttk.Treeview(parent, columns=cols, show="headings",
                            style="CAT.Treeview", selectmode="browse")

        headers = ("Название", "Автор", "Год", "Издательство", "Категория",
                   "Имя файла", "Формат", "Размер МБ", "Статус")
        widths  = (240,        150,     55,    130,           160,
                   190,        60,      80,    75)

        for col, hdr, w in zip(cols, headers, widths):
            tree.heading(col, text=hdr, anchor="w",
                         command=lambda c=col: self._sort_by(c))
            tree.column(col, width=w, minwidth=40, stretch=False)

        tree.tag_configure("tag_cache", background="#1a3a1a")
        tree.tag_configure("tag_ai",    background="#1a2a3a")
        tree.tag_configure("tag_error", background="#3a1a1a")
        tree.tag_configure("tag_media", background="#1a2a1a", foreground="#80d080")

        vsb = ttk.Scrollbar(parent, orient="vertical",   command=tree.yview)
        hsb = ttk.Scrollbar(parent, orient="horizontal", command=tree.xview)
        tree.configure(yscrollcommand=vsb.set, xscrollcommand=hsb.set)
        vsb.pack(side="right",  fill="y")
        hsb.pack(side="bottom", fill="x")
        tree.pack(fill="both", expand=True)

        tree.bind("<<TreeviewSelect>>", self._on_select)
        tree.bind("<Double-Button-1>", self._on_double_click)
        tree.bind("<Button-3>", self._on_right_click)

        self._ctx_menu = tk.Menu(tree, tearoff=0)
        self._ctx_menu.add_command(label="Открыть файл", command=self._open_file)
        self._ctx_menu.add_command(label="Показать в проводнике", command=self._open_folder)
        self._ctx_menu.add_separator()
        self._ctx_menu.add_command(label="Копировать путь", command=self._copy_path)

        return tree

    def _build_right_panel(self, parent):
        ctk.CTkLabel(parent, text="Редактирование записи",
                     font=ctk.CTkFont(size=11, weight="bold"),
                     anchor="w").pack(fill="x", padx=10, pady=(8, 4))

        def field(label: str) -> tk.StringVar:
            row = ctk.CTkFrame(parent, fg_color="transparent")
            row.pack(fill="x", padx=10, pady=(0, 3))
            ctk.CTkLabel(row, text=label, anchor="w", width=72).pack(side="left")
            var = tk.StringVar()
            ctk.CTkEntry(row, textvariable=var).pack(side="left", fill="x", expand=True)
            return var

        self._ev_title     = field("Название:")
        self._ev_author    = field("Автор:")
        self._ev_year      = field("Год:")
        self._ev_publisher = field("Изд-во:")

        self._apply_btn = ctk.CTkButton(
            parent, text="Применить метаданные", state="disabled",
            fg_color="#1a6b3a", hover_color="#145530",
            command=self._apply_edit)
        self._apply_btn.pack(fill="x", padx=10, pady=(4, 4))

        # file path
        self._path_label = ctk.CTkLabel(
            parent, text="", anchor="w", text_color="gray50",
            font=ctk.CTkFont(size=9), wraplength=335)
        self._path_label.pack(fill="x", padx=10, pady=(0, 3))

        # summary from AI
        ctk.CTkLabel(parent, text="Саммари:",
                     font=ctk.CTkFont(size=9), text_color="gray55",
                     anchor="w").pack(fill="x", padx=10, pady=(2, 0))
        self._desc_box = ctk.CTkTextbox(
            parent, height=130, fg_color="#1a1a2a", border_width=0,
            font=ctk.CTkFont(family="Segoe UI", size=12), wrap="word", state="disabled")
        self._desc_box.pack(fill="x", padx=10, pady=(2, 6))

        # divider
        ctk.CTkFrame(parent, height=1, fg_color="gray35").pack(fill="x", padx=10, pady=(0, 6))

        # current category
        ctk.CTkLabel(parent, text="Текущая категория:",
                     font=ctk.CTkFont(size=10, weight="bold"),
                     anchor="w").pack(fill="x", padx=10)
        self._cat_label = ctk.CTkLabel(
            parent, text="—", anchor="w",
            text_color="#7eb8f7", wraplength=330,
            font=ctk.CTkFont(size=11))
        self._cat_label.pack(fill="x", padx=10, pady=(2, 6))

        # two-column picker
        ctk.CTkLabel(parent, text="Выбрать категорию:",
                     font=ctk.CTkFont(size=10, weight="bold"),
                     anchor="w").pack(fill="x", padx=10)

        picker = ctk.CTkFrame(parent, fg_color="transparent")
        picker.pack(fill="both", expand=True, padx=10, pady=(2, 4))

        lb_kw = dict(bg="#1e1e2e", fg="#dde0e6",
                     selectbackground="#1a5276", selectforeground="white",
                     font=("Segoe UI", 9), borderwidth=1,
                     highlightthickness=0, activestyle="none",
                     exportselection=False)

        # left column — sections
        sec_col = ctk.CTkFrame(picker, fg_color="transparent")
        sec_col.pack(side="left", fill="both", expand=True, padx=(0, 3))
        ctk.CTkLabel(sec_col, text="Раздел", anchor="w",
                     font=ctk.CTkFont(size=9), text_color="gray55").pack(fill="x")
        sec_vsb = ttk.Scrollbar(sec_col, orient="vertical")
        self._section_lb = tk.Listbox(sec_col, **lb_kw,
                                      yscrollcommand=sec_vsb.set)
        sec_vsb.configure(command=self._section_lb.yview)
        sec_vsb.pack(side="right", fill="y")
        self._section_lb.pack(side="left", fill="both", expand=True)

        # right column — subcategories
        sub_col = ctk.CTkFrame(picker, fg_color="transparent")
        sub_col.pack(side="right", fill="both", expand=True)
        ctk.CTkLabel(sub_col, text="Подраздел", anchor="w",
                     font=ctk.CTkFont(size=9), text_color="gray55").pack(fill="x")
        sub_vsb = ttk.Scrollbar(sub_col, orient="vertical")
        self._sub_lb = tk.Listbox(sub_col, **lb_kw,
                                  yscrollcommand=sub_vsb.set)
        sub_vsb.configure(command=self._sub_lb.yview)
        sub_vsb.pack(side="right", fill="y")
        self._sub_lb.pack(side="left", fill="both", expand=True)

        # populate sections
        for sec in sorted(self._taxonomy):
            self._section_lb.insert("end", sec)

        self._section_lb.bind("<<ListboxSelect>>", self._on_section_select)
        self._sub_lb.bind("<<ListboxSelect>>",     self._on_sub_select)

        # free-text fallback
        ctk.CTkFrame(parent, height=1, fg_color="gray35").pack(fill="x", padx=10, pady=(0, 4))
        ctk.CTkLabel(parent, text="Свой вариант:", anchor="w",
                     font=ctk.CTkFont(size=10)).pack(fill="x", padx=10)
        free_row = ctk.CTkFrame(parent, fg_color="transparent")
        free_row.pack(fill="x", padx=10, pady=(2, 10))
        self._free_cat_var = tk.StringVar()
        ctk.CTkEntry(free_row, textvariable=self._free_cat_var,
                     placeholder_text="Раздел/Подраздел").pack(
                         side="left", fill="x", expand=True, padx=(0, 4))
        ctk.CTkButton(free_row, text="Ок", width=44,
                      command=self._apply_free_category).pack(side="right")

    @staticmethod
    def _build_taxonomy(categories: list[str]) -> dict[str, list[str]]:
        taxonomy: dict[str, list[str]] = {}
        for cat in categories:
            if "/" in cat:
                section, sub = cat.split("/", 1)
                taxonomy.setdefault(section, []).append(sub)
            else:
                taxonomy.setdefault(cat, [])
        return taxonomy

    def reload_taxonomy(self) -> None:
        """Refresh the section/subcategory picker from the active taxonomy."""
        self._taxonomy = self._build_taxonomy(_get_taxonomy()[0])
        self._section_lb.delete(0, "end")
        self._sub_lb.delete(0, "end")
        for sec in sorted(self._taxonomy):
            self._section_lb.insert("end", sec)

    # --------------------------------------------------------------- load

    def _load_from_scan(self):
        files = self._get_files()
        if not files:
            self._stats_label.configure(
                text="Нет файлов — сначала выполните сканирование")
            return
        root_dir = self._get_scan_root() if self._get_scan_root else ""
        cache_utils.save_file_list(
            [f for f in files if not getattr(f, "is_media_dir", False)],
            root_dir=root_dir,
        )
        self._populate(files)

    def _load_from_disk(self):
        files, saved_at, _ = cache_utils.load_file_list()
        if not files:
            self._stats_label.configure(
                text="Сохранённый список не найден — выполните сканирование")
            return
        date_str = ""
        if saved_at:
            try:
                from datetime import datetime
                dt = datetime.fromisoformat(saved_at)
                date_str = f"  (сохранено {dt.strftime('%d.%m.%Y %H:%M')})"
            except Exception:
                pass
        self._populate(files, note=date_str)

    def _populate(self, files: list[FileInfo], note: str = ""):
        self._cache = cache_utils.load()
        self._entries.clear()

        cached_count = 0
        for file_info in files:
            hit = cache_utils.get_cached(file_info, self._cache)
            if hit:
                entry = self._make_entry(file_info, hit,
                                         "кэш" if "error" not in hit else "ошибка")
                cached_count += 1
            else:
                entry = self._make_entry(file_info, {}, "")
            self._entries.append(entry)

        need = len(files) - cached_count
        self._stats_label.configure(
            text=f"{len(files)} файлов  |  кэш: {cached_count}  |  ожидают ИИ: {need}{note}")
        self._reset_btn.configure(state="normal" if files else "disabled")
        self._export_btn.configure(state="normal" if files else "disabled")
        self._update_btn.configure(state="normal" if files else "disabled")
        self._update_filter_options()
        self._apply_filter()
        self._update_classify_button()

        if need == 0 and cached_count > 0:
            self._status_label.configure(
                text="Все файлы уже классифицированы. "
                     "Нажмите «Сбросить классификацию» для повторной обработки.")

    def _reset_classifications(self):
        self._cache = {}
        cache_utils.save(self._cache)
        for entry in self._entries:
            entry.update({"title": "", "author": "", "year": "",
                          "publisher": "", "category": "", "status": ""})
        n = len(self._entries)
        self._stats_label.configure(
            text=f"{n} файлов  |  кэш: 0  |  ожидают ИИ: {n}")
        self._progress.set(0)
        self._status_label.configure(text="Классификация сброшена. Нажмите «Классифицировать».")
        self._filter_var.set("Все")
        self._update_filter_options()
        self._apply_filter()
        self._update_classify_button()

    # ----------------------------------------------------------- classify

    def _update_classify_button(self):
        if not self._entries:
            self._classify_btn.configure(state="disabled")
            return
        metadata_only = self._meta_only_var.get()
        if metadata_only:
            can = any(e["status"] in ("", "ошибка") for e in self._entries)
        else:
            can = any(
                e["status"] in ("", "ошибка") or (e["status"] and not e["category"])
                for e in self._entries
            )
        self._classify_btn.configure(state="normal" if can else "disabled")

    def _start_classify(self):
        metadata_only = self._meta_only_var.get()

        for e in self._entries:
            if e["status"] == "ошибка":
                e["status"] = ""
        if not metadata_only:
            for e in self._entries:
                if e["status"] and not e["category"]:
                    e["status"] = ""

        to_do = [e for e in self._entries if e["status"] == ""]
        if not to_do:
            self._status_label.configure(
                text="Нет файлов для классификации. Сбросьте классификацию для повторной обработки.")
            return

        self._classify_gen += 1
        gen = self._classify_gen
        self._classifier = Classifier()
        self._classify_btn.configure(state="disabled")
        self._stop_btn.configure(state="normal")
        self._progress.set(0)
        if not metadata_only:
            self._apply_filter()
        threading.Thread(
            target=self._run_classify, args=(to_do, gen, metadata_only), daemon=True
        ).start()

    def _run_classify(self, to_do: list[dict], gen: int, metadata_only: bool):
        files = [e["file"] for e in to_do]
        path_to_idx = {e["file"].full_path: i
                       for i, e in enumerate(self._entries)}
        self._classifier.classify_all(
            files=files,
            cache=self._cache,
            on_result=lambda i, info, res, from_cache:
                self._queue.put(("result", gen, path_to_idx[info.full_path], res, from_cache)),
            on_progress=lambda cur, tot:
                self._queue.put(("progress", gen, cur, tot)),
            on_done=lambda _cache:
                self._queue.put(("done", gen)),
            on_error=lambda fp, err:
                self._queue.put(("err", gen, fp, err)),
            metadata_only=metadata_only,
        )

    def _stop_classify(self):
        if self._classifier:
            self._classifier.stop()
        self._stop_btn.configure(state="disabled")

    # ----------------------------------------------------------- queue pump

    def _poll_queue(self):
        try:
            for _ in range(30):
                msg = self._queue.get_nowait()
                kind = msg[0]

                if kind == "progress":
                    _, gen, cur, tot = msg
                    if gen != self._classify_gen:
                        continue
                    self._progress.set(cur / tot if tot else 0)
                    errors = sum(1 for e in self._entries if e["status"] == "ошибка")
                    err_str = f"  |  ошибок: {errors}" if errors else ""
                    self._status_label.configure(
                        text=f"Обрабатывается: {cur} / {tot}{err_str}")

                elif kind == "result":
                    _, gen, idx, result, from_cache = msg
                    if gen != self._classify_gen:
                        continue
                    entry = self._entries[idx]
                    self._fill_entry_from_result(entry, result, from_cache)
                    self._update_row(idx)
                    done = sum(1 for e in self._entries if e["status"])
                    total = len(self._entries)
                    self._stats_label.configure(
                        text=f"{total} файлов  |  классифицировано: {done}")

                elif kind == "done":
                    _, gen = msg
                    if gen != self._classify_gen:
                        continue
                    self._progress.set(1.0)
                    done = sum(1 for e in self._entries if e["status"])
                    errors = sum(1 for e in self._entries if e["status"] == "ошибка")
                    no_cat = sum(1 for e in self._entries if e["status"] and not e["category"])
                    self._status_label.configure(
                        text=f"Завершено. Классифицировано: {done}"
                        + (f"  |  ошибок: {errors}" if errors else "")
                        + (f"  |  без категории: {no_cat} (запустите полную классификацию)" if no_cat else ""))
                    self._stop_btn.configure(state="disabled")
                    self._export_btn.configure(state="normal")
                    self._update_filter_options()
                    self._apply_filter()
                    self._update_classify_button()

                elif kind == "err":
                    pass

                elif kind == "incremental_done":
                    _, new_files, root_dir = msg
                    self._apply_incremental_result(new_files, root_dir)

        except queue.Empty:
            pass
        finally:
            self.after(150, self._poll_queue)

    # -------------------------------------------------- incremental update

    def _start_incremental_update(self):
        """Quick scan to detect added/removed files without reclassifying."""
        _, _, root_dir = cache_utils.load_file_list()
        # Only fall back to scan tab if the saved root_dir is known and non-empty
        if not root_dir:
            self._status_label.configure(
                text="Корневой каталог не сохранён. "
                     "Выполните сканирование и нажмите «Загрузить из сканирования».",
                text_color="#f44336")
            return
        if not self._entries:
            self._status_label.configure(
                text="Список пуст. Сначала загрузите каталог.",
                text_color="gray60")
            return

        self._update_btn.configure(state="disabled")
        self._classify_btn.configure(state="disabled")
        self._progress.set(0)
        self._status_label.configure(
            text=f"Поиск изменений в {root_dir} …", text_color="gray60")

        threading.Thread(
            target=self._run_incremental_update,
            args=(root_dir,), daemon=True,
        ).start()

    def _run_incremental_update(self, root_dir: str):
        from core.scanner import Scanner
        scanner = Scanner()
        new_files = scanner.scan(root_dir, compute_hashes=False)
        self._queue.put(("incremental_done", new_files, root_dir))

    def _apply_incremental_result(self, new_files: list, root_dir: str):
        current_paths = {e["file"].full_path for e in self._entries
                         if not e["file"].is_media_dir}

        # Safety: if scan returned nothing, abort — likely wrong/inaccessible path
        if not new_files and current_paths:
            self._progress.set(0)
            self._status_label.configure(
                text=f"Ошибка: каталог '{root_dir}' недоступен или не содержит файлов. "
                     f"Изменения не применены.",
                text_color="#f44336")
            self._update_btn.configure(state="normal")
            self._update_classify_button()
            return

        new_paths = {f.full_path for f in new_files}
        paths_to_remove = current_paths - new_paths

        # Safety: abort if >80% of known files would be removed (wrong path likely)
        if current_paths and len(paths_to_remove) > len(current_paths) * 0.8:
            self._progress.set(0)
            self._status_label.configure(
                text=f"Ошибка: {len(paths_to_remove)} из {len(current_paths)} файлов не найдено. "
                     f"Проверьте доступность каталога '{root_dir}'. Изменения не применены.",
                text_color="#f44336")
            self._update_btn.configure(state="normal")
            self._update_classify_button()
            return

        files_to_add = [f for f in new_files if f.full_path not in current_paths]

        # Remove entries for deleted files (keep media dirs)
        self._entries = [
            e for e in self._entries
            if e["file"].is_media_dir or e["file"].full_path not in paths_to_remove
        ]

        # Add new entries (unclassified or from cache)
        for f in files_to_add:
            cached = cache_utils.get_cached(f, self._cache)
            if cached:
                entry = self._make_entry(f, cached,
                                          "кэш" if "error" not in cached else "ошибка")
            else:
                entry = self._make_entry(f, {}, "")
            self._entries.append(entry)

        # Save updated file list (doc files only, preserving SHA256 for existing files)
        doc_files = [e["file"] for e in self._entries if not e["file"].is_media_dir]
        cache_utils.save_file_list(doc_files, root_dir)

        n_added = len(files_to_add)
        n_removed = len(paths_to_remove)
        total = len(self._entries)
        need = sum(1 for e in self._entries if e["status"] == "")

        self._progress.set(1.0)
        self._stats_label.configure(
            text=f"{total} файлов  |  кэш: {total - need}  |  ожидают ИИ: {need}")
        self._status_label.configure(
            text=f"Обновление завершено. Добавлено: {n_added}, удалено: {n_removed}.",
            text_color="gray60")
        self._update_btn.configure(state="normal")
        self._update_filter_options()
        self._apply_filter()
        self._update_classify_button()

    # ----------------------------------------------------------- table

    def _insert_row(self, entry: dict) -> str:
        return self._tree.insert(
            "", "end",
            values=self._row_values(entry),
            tags=(self._row_tag(entry),),
        )

    def _update_row(self, idx: int):
        iid = self._idx_to_iid.get(idx)
        if iid:
            entry = self._entries[idx]
            self._tree.item(iid, values=self._row_values(entry),
                            tags=(self._row_tag(entry),))

    @staticmethod
    def _row_values(e: dict) -> tuple:
        f: FileInfo | MediaDirInfo = e["file"]
        name = f"📁  {f.name}" if f.is_media_dir else f.name
        return (
            e["title"] or "—",
            e["author"] or "—",
            e["year"] or "—",
            e["publisher"] or "—",
            e["category"] or "—",
            name,
            f.extension,
            f"{f.size_mb:.2f}",
            e["status"],
        )

    @staticmethod
    def _row_tag(e: dict) -> str:
        if e["file"].is_media_dir:
            return "tag_media"
        return {"кэш": "tag_cache", "ИИ": "tag_ai",
                "ошибка": "tag_error"}.get(e["status"], "")

    @staticmethod
    def _make_entry(file_info: FileInfo, result: dict, status: str) -> dict:
        return {
            "file":      file_info,
            "title":     (result.get("title") or "").strip(),
            "author":    (result.get("author") or "").strip(),
            "year":      str(result.get("year") or "").strip(),
            "publisher": (result.get("publisher") or "").strip(),
            "category":  result.get("category") or "",
            "status":    status,
        }

    @staticmethod
    def _fill_entry_from_result(entry: dict, result: dict, from_cache: bool):
        entry["title"]     = (result.get("title") or "").strip()
        entry["author"]    = (result.get("author") or "").strip()
        entry["year"]      = str(result.get("year") or "").strip()
        entry["publisher"] = (result.get("publisher") or "").strip()
        if "category" in result:
            entry["category"] = result["category"] or "Прочее"
        entry["status"]    = "кэш" if from_cache else ("ошибка" if "error" in result else "ИИ")

    # ----------------------------------------------------------- filter / sort

    def _apply_filter(self):
        category = self._filter_var.get()
        search = self._search_var.get().lower().strip()

        for item in self._tree.get_children():
            self._tree.delete(item)
        self._iid_to_idx.clear()
        self._idx_to_iid.clear()

        for idx, entry in enumerate(self._entries):
            if category not in ("Все",):
                cat = entry["category"]
                if cat != category and not cat.startswith(category + "/"):
                    continue
            if search:
                haystack = " ".join([
                    entry["title"], entry["author"], entry["file"].name
                ]).lower()
                if search not in haystack:
                    continue
            iid = self._insert_row(entry)
            self._iid_to_idx[iid] = idx
            self._idx_to_iid[idx] = iid

    def _update_filter_options(self):
        cats = sorted({e["category"] for e in self._entries if e["category"]})
        sections = sorted({c.split("/")[0] for c in cats if "/" in c})
        combined = ["Все"] + sections
        for c in cats:
            if c not in combined:
                combined.append(c)
        self._filter_box.configure(values=combined)

    _sort_rev: dict[str, bool] = {}

    def _sort_by(self, col: str):
        rev = self._sort_rev.get(col, False)
        items = [(self._tree.set(k, col), k) for k in self._tree.get_children("")]
        try:
            items.sort(key=lambda x: float(x[0].replace("—", "0")), reverse=rev)
        except ValueError:
            items.sort(key=lambda x: x[0].lower(), reverse=rev)
        for i, (_, k) in enumerate(items):
            self._tree.move(k, "", i)
        self._sort_rev[col] = not rev

    # ----------------------------------------------------------- right panel logic

    def _on_select(self, _event):
        sel = self._tree.selection()
        if not sel:
            self._apply_btn.configure(state="disabled")
            self._path_label.configure(text="")
            self._set_description("")
            return
        iid = sel[0]
        idx = self._iid_to_idx.get(iid, -1)
        if idx < 0:
            return
        self._edit_idx = idx
        e = self._entries[idx]
        self._ev_title.set(e["title"])
        self._ev_author.set(e["author"])
        self._ev_year.set(e["year"])
        self._ev_publisher.set(e["publisher"])
        self._cat_label.configure(text=e["category"] or "—")
        self._apply_btn.configure(state="normal")
        self._path_label.configure(text=e["file"].full_path)
        self._preselect_category(e["category"])

        if e["file"].is_media_dir:
            md: MediaDirInfo = e["file"]
            subs = ", ".join(md.subdirs[:8]) or "—"
            self._set_description(
                f"Медиа-каталог\n"
                f"Медиафайлов: {md.media_count}\n"
                f"Размер: {md.size_mb:.1f} МБ\n"
                f"Подпапки: {subs}"
            )
        else:
            sum_cache = cache_utils.load_summaries()
            summary = cache_utils.get_summary(e["file"], sum_cache)
            self._set_description(summary or "")

        if e["status"] == "ошибка":
            cached = self._cache.get(cache_utils._key(e["file"]), {})
            err_msg = cached.get("error", "причина неизвестна")
            self._status_label.configure(
                text=f"Ошибка «{e['file'].name}»: {err_msg}",
                text_color="#f44336")
        else:
            self._status_label.configure(text_color="gray60")

    def _preselect_category(self, category: str):
        self._preselecting = True
        try:
            self._section_lb.selection_clear(0, "end")
            self._sub_lb.delete(0, "end")
            if not category:
                return
            section, sub = (category.split("/", 1) + [None])[:2] if "/" in category \
                           else (category, None)
            sections = list(self._section_lb.get(0, "end"))
            if section not in sections:
                return
            si = sections.index(section)
            self._section_lb.selection_set(si)
            self._section_lb.see(si)
            subs = self._taxonomy.get(section, [])
            for s in subs:
                self._sub_lb.insert("end", s)
            if sub and sub in subs:
                subi = subs.index(sub)
                self._sub_lb.selection_set(subi)
                self._sub_lb.see(subi)
        finally:
            self._preselecting = False

    def _on_section_select(self, _event):
        if self._preselecting:
            return
        sel = self._section_lb.curselection()
        if not sel:
            return
        section = self._section_lb.get(sel[0])
        subs = self._taxonomy.get(section, [])
        self._sub_lb.delete(0, "end")
        if not subs:
            # leaf section (e.g. "Прочее", "Химия") — apply immediately
            self._apply_category(section)
        else:
            for s in subs:
                self._sub_lb.insert("end", s)

    def _on_sub_select(self, _event):
        if self._preselecting:
            return
        sel_sec = self._section_lb.curselection()
        sel_sub = self._sub_lb.curselection()
        if not sel_sec or not sel_sub:
            return
        section = self._section_lb.get(sel_sec[0])
        sub     = self._sub_lb.get(sel_sub[0])
        self._apply_category(f"{section}/{sub}")

    def _apply_category(self, category: str):
        idx = self._edit_idx
        if idx < 0 or idx >= len(self._entries):
            return
        e = self._entries[idx]
        e["category"] = category
        self._cat_label.configure(text=category)
        cache_utils.set_cached(e["file"], {
            "title": e["title"], "author": e["author"],
            "year": e["year"], "publisher": e["publisher"],
            "category": category,
        }, self._cache)
        cache_utils.save(self._cache)
        self._update_row(idx)
        self._update_filter_options()

    def _apply_free_category(self):
        cat = self._free_cat_var.get().strip()
        if cat:
            self._apply_category(cat)
            self._free_cat_var.set("")

    # ----------------------------------------------------------- file actions

    def _current_path(self) -> str | None:
        if self._edit_idx < 0 or self._edit_idx >= len(self._entries):
            return None
        return self._entries[self._edit_idx]["file"].full_path

    def _open_file(self):
        path = self._current_path()
        if path:
            try:
                os.startfile(path)
            except Exception:
                pass

    def _open_folder(self):
        path = self._current_path()
        if path:
            try:
                subprocess.Popen(f'explorer /select,"{os.path.normpath(path)}"')
            except Exception:
                pass

    def _copy_path(self):
        path = self._current_path()
        if path:
            self.clipboard_clear()
            self.clipboard_append(path)

    def _on_double_click(self, event):
        if self._tree.identify_row(event.y):
            self._open_file()

    def _on_right_click(self, event):
        item = self._tree.identify_row(event.y)
        if item:
            self._tree.selection_set(item)
            self._ctx_menu.post(event.x_root, event.y_root)

    # ----------------------------------------------------------- description

    def _set_description(self, text: str):
        self._desc_box.configure(state="normal")
        self._desc_box.delete("1.0", "end")
        self._desc_box.insert("1.0", text if text else "(саммари не составлено — перейдите на вкладку «Саммари»)")
        self._desc_box.configure(state="disabled")

    def _apply_edit(self):
        idx = self._edit_idx
        if idx < 0 or idx >= len(self._entries):
            return
        e = self._entries[idx]
        e["title"]     = self._ev_title.get().strip()
        e["author"]    = self._ev_author.get().strip()
        e["year"]      = self._ev_year.get().strip()
        e["publisher"] = self._ev_publisher.get().strip()
        cache_utils.set_cached(e["file"], {
            "title": e["title"], "author": e["author"],
            "year": e["year"], "publisher": e["publisher"],
            "category": e["category"],
        }, self._cache)
        cache_utils.save(self._cache)
        self._update_row(idx)
        self._update_filter_options()

    # ----------------------------------------------------------- excel export

    def _export_excel(self):
        if not self._entries:
            return
        from tkinter import filedialog
        path = filedialog.asksaveasfilename(
            defaultextension=".xlsx",
            filetypes=[("Excel файл", "*.xlsx")],
            initialfile="catalog.xlsx",
            title="Сохранить каталог как…",
        )
        if not path:
            return
        try:
            self._write_excel(path)
            self._status_label.configure(
                text=f"Экспорт завершён: {path}", text_color="gray60")
            try:
                os.startfile(path)
            except Exception:
                pass
        except Exception as exc:
            self._status_label.configure(
                text=f"Ошибка экспорта: {exc}", text_color="#f44336")

    def _write_excel(self, path: str):
        from openpyxl import Workbook
        from openpyxl.styles import (Font, PatternFill, Alignment,
                                     Border, Side)
        from openpyxl.utils import get_column_letter

        sum_cache = cache_utils.load_summaries()

        wb = Workbook()
        ws = wb.active
        ws.title = "Каталог"

        # ---- styles ----
        header_font   = Font(name="Calibri", bold=True, color="FFFFFF", size=11)
        header_fill   = PatternFill("solid", fgColor="1F538D")
        header_align  = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell_align    = Alignment(vertical="top", wrap_text=False)
        wrap_align    = Alignment(vertical="top", wrap_text=True)
        thin          = Side(style="thin", color="C0C0C0")
        border        = Border(left=thin, right=thin, top=thin, bottom=thin)

        alt_fill   = PatternFill("solid", fgColor="F7FBFF")
        white_fill = PatternFill("solid", fgColor="FFFFFF")
        err_fill   = PatternFill("solid", fgColor="FDECEA")

        link_font  = Font(name="Calibri", color="1155CC", underline="single", size=10)
        base_font  = Font(name="Calibri", size=10)

        # ---- header row ----
        headers = [
            "№", "Название", "Автор", "Год", "Издательство",
            "Категория", "Раздел", "Подраздел",
            "Формат", "Размер МБ", "Статус",
            "Имя файла", "Папка", "Ссылка на файл", "Саммари",
        ]
        LINK_COL    = 14   # "Ссылка на файл"
        SUMMARY_COL = 15   # "Саммари"

        ws.append(headers)
        for col, _ in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col)
            cell.font      = header_font
            cell.fill      = header_fill
            cell.alignment = header_align
            cell.border    = border
        ws.row_dimensions[1].height = 30

        # ---- data rows ----
        for row_idx, e in enumerate(self._entries, 2):
            f = e["file"]
            cat = e["category"] or ""
            section, sub = (cat.split("/", 1) if "/" in cat else (cat, ""))

            summary = ""
            if not f.is_media_dir:
                s = cache_utils.get_summary(f, sum_cache)
                summary = s or ""

            # Use HYPERLINK formula — path stored as UTF-8 string, works with Cyrillic UNC paths
            safe_path = f.full_path.replace('"', "'")
            safe_name = f.name.replace('"', "'")
            hyperlink_formula = f'=HYPERLINK("{safe_path}","{safe_name}")'

            row_data = [
                row_idx - 1,
                e["title"] or "",
                e["author"] or "",
                e["year"] or "",
                e["publisher"] or "",
                cat,
                section,
                sub,
                f.extension,
                round(f.size_mb, 2),
                e["status"],
                f.name,
                f.directory,
                hyperlink_formula,
                summary,
            ]
            ws.append(row_data)

            # determine row fill
            if e["status"] == "ошибка":
                row_fill = err_fill
            elif row_idx % 2 == 0:
                row_fill = white_fill
            else:
                row_fill = alt_fill

            for col, _ in enumerate(row_data, 1):
                cell = ws.cell(row=row_idx, column=col)
                cell.fill      = row_fill
                cell.border    = border
                cell.font      = base_font
                cell.alignment = wrap_align if col == SUMMARY_COL else cell_align

            # style the HYPERLINK formula cell
            link_cell = ws.cell(row=row_idx, column=LINK_COL)
            link_cell.font = link_font

        # ---- freeze header, auto-filter ----
        ws.freeze_panes = "A2"
        ws.auto_filter.ref = ws.dimensions

        # ---- column widths (sample first 200 rows + header) ----
        col_widths = [len(h) for h in headers]
        for e in self._entries[:200]:
            f = e["file"]
            cat = e["category"] or ""
            section, sub = (cat.split("/", 1) if "/" in cat else (cat, ""))
            s = cache_utils.get_summary(f, sum_cache) or "" if not f.is_media_dir else ""
            vals = [
                len(str(self._entries.index(e) + 1)),
                len(e["title"] or ""),
                len(e["author"] or ""),
                len(e["year"] or ""),
                len(e["publisher"] or ""),
                len(cat),
                len(section),
                len(sub),
                len(f.extension),
                5,
                len(e["status"]),
                len(f.name),
                len(f.directory),
                len(f.name),
                min(len(s), 60),
            ]
            col_widths = [max(a, b) for a, b in zip(col_widths, vals)]

        max_widths = [6, 50, 35, 6, 30, 45, 30, 30, 8, 10, 10, 50, 60, 50, 80]
        for col, (w, mx) in enumerate(zip(col_widths, max_widths), 1):
            ws.column_dimensions[get_column_letter(col)].width = min(w + 2, mx)

        wb.save(path)
