import os
import queue
import threading
import tkinter as tk
from pathlib import Path
from tkinter import ttk, filedialog
from typing import Callable

import customtkinter as ctk

from core.scanner import FileInfo
from core.summarizer import Summarizer
from utils import cache as cache_utils


class SummaryTab(ctk.CTkFrame):
    def __init__(self, parent, get_files: Callable[[], list[FileInfo]]):
        super().__init__(parent, fg_color="transparent")
        self._get_files = get_files
        self._entries: list[dict] = []
        self._iid_to_idx: dict[str, int] = {}
        self._idx_to_iid: dict[int, str] = {}
        self._sum_cache: dict = {}
        self._summarizer: Summarizer | None = None
        self._queue: queue.Queue = queue.Queue()
        self._gen: int = 0
        self._selected_idx: int = -1

        self._build_ui()
        self._poll_queue()

    # ---------------------------------------------------------------- build ui

    def _build_ui(self):
        # Row 1: load buttons + stats
        top = ctk.CTkFrame(self, fg_color="transparent")
        top.pack(fill="x", padx=12, pady=(10, 2))

        ctk.CTkButton(top, text="Загрузить из сканирования", width=215,
                      command=self._load_from_scan).pack(side="left", padx=(0, 8))
        ctk.CTkButton(top, text="Загрузить с диска", width=150,
                      fg_color="gray30", hover_color="gray25",
                      command=self._load_from_disk).pack(side="left", padx=(0, 8))

        self._export_btn = ctk.CTkButton(
            top, text="Экспорт в Excel", width=150, state="disabled",
            fg_color="gray30", hover_color="gray25",
            command=self._export_excel)
        self._export_btn.pack(side="right")

        self._stats_label = ctk.CTkLabel(top, text="Файлы не загружены",
                                          text_color="gray60")
        self._stats_label.pack(side="right", padx=12)

        # Row 2: action buttons
        top2 = ctk.CTkFrame(self, fg_color="transparent")
        top2.pack(fill="x", padx=12, pady=(0, 4))

        self._run_btn = ctk.CTkButton(
            top2, text="Составить саммари", width=180, state="disabled",
            command=self._start_summarize)
        self._run_btn.pack(side="left", padx=(0, 8))

        self._stop_btn = ctk.CTkButton(
            top2, text="Стоп", width=80, state="disabled",
            fg_color="#555", hover_color="#444",
            command=self._stop_summarize)
        self._stop_btn.pack(side="left", padx=(0, 16))

        self._reset_btn = ctk.CTkButton(
            top2, text="Сбросить саммари", width=180, state="disabled",
            fg_color="#6b1a1a", hover_color="#551515",
            command=self._reset_summaries)
        self._reset_btn.pack(side="left")

        # Progress
        prog = ctk.CTkFrame(self, fg_color="transparent")
        prog.pack(fill="x", padx=12, pady=(0, 4))

        self._progress = ctk.CTkProgressBar(prog)
        self._progress.pack(fill="x")
        self._progress.set(0)

        self._status_label = ctk.CTkLabel(
            prog, text="", text_color="gray60",
            anchor="w", font=ctk.CTkFont(size=11))
        self._status_label.pack(fill="x", pady=(2, 0))

        # Main area: table + right panel
        main_area = ctk.CTkFrame(self, fg_color="transparent")
        main_area.pack(fill="both", expand=True, padx=12, pady=(0, 10))

        table_frame = ctk.CTkFrame(main_area)
        table_frame.pack(side="left", fill="both", expand=True)
        self._tree = self._create_tree(table_frame)

        right = ctk.CTkFrame(main_area, width=350)
        right.pack(side="right", fill="y", padx=(8, 0))
        right.pack_propagate(False)
        self._build_right_panel(right)

    def _create_tree(self, parent) -> ttk.Treeview:
        style = ttk.Style()
        bg, fg, sel, head = "#2b2b2b", "#e0e0e0", "#1a5276", "#1f538d"

        style.configure("SUM.Treeview",
                        background=bg, foreground=fg, fieldbackground=bg,
                        rowheight=22, borderwidth=0, font=("Segoe UI", 10))
        style.configure("SUM.Treeview.Heading",
                        background=head, foreground="white",
                        font=("Segoe UI", 10, "bold"), relief="flat")
        style.map("SUM.Treeview", background=[("selected", sel)])
        style.map("SUM.Treeview.Heading", background=[("active", "#144578")])

        cols = ("title", "author", "category", "name", "ext", "summary_preview", "status")
        tree = ttk.Treeview(parent, columns=cols, show="headings",
                            style="SUM.Treeview", selectmode="browse")

        headers = ("Название", "Автор", "Категория", "Имя файла", "Формат", "Саммари", "Статус")
        widths  = (210,        130,     170,         180,          55,       300,        60)

        for col, hdr, w in zip(cols, headers, widths):
            tree.heading(col, text=hdr, anchor="w",
                         command=lambda c=col: self._sort_by(c))
            tree.column(col, width=w, minwidth=40,
                        stretch=(col == "summary_preview"))

        tree.tag_configure("tag_cache", background="#1a3a1a")
        tree.tag_configure("tag_ai",    background="#1a2a3a")
        tree.tag_configure("tag_error", background="#3a1a1a")

        vsb = ttk.Scrollbar(parent, orient="vertical",   command=tree.yview)
        hsb = ttk.Scrollbar(parent, orient="horizontal", command=tree.xview)
        tree.configure(yscrollcommand=vsb.set, xscrollcommand=hsb.set)
        vsb.pack(side="right",  fill="y")
        hsb.pack(side="bottom", fill="x")
        tree.pack(fill="both", expand=True)

        tree.bind("<<TreeviewSelect>>", self._on_select)
        return tree

    def _build_right_panel(self, parent):
        ctk.CTkLabel(parent, text="Полное саммари",
                     font=ctk.CTkFont(size=11, weight="bold"),
                     anchor="w").pack(fill="x", padx=10, pady=(8, 4))

        self._file_label = ctk.CTkLabel(
            parent, text="", anchor="w", text_color="gray50",
            font=ctk.CTkFont(size=9), wraplength=330)
        self._file_label.pack(fill="x", padx=10, pady=(0, 6))

        self._summary_box = ctk.CTkTextbox(
            parent, fg_color="#1a1a2a", border_width=0,
            font=ctk.CTkFont(family="Segoe UI", size=12),
            wrap="word", state="disabled")
        self._summary_box.pack(fill="both", expand=True, padx=10, pady=(0, 10))

    # ---------------------------------------------------------------- load

    def _load_from_scan(self):
        files = self._get_files()
        if not files:
            self._stats_label.configure(
                text="Нет файлов — сначала выполните сканирование")
            return
        self._populate(files)

    def _load_from_disk(self):
        files, _ = cache_utils.load_file_list()
        if not files:
            self._stats_label.configure(
                text="Сохранённый список не найден — выполните сканирование")
            return
        self._populate(files)

    def _populate(self, files: list[FileInfo]):
        cls_cache = cache_utils.load()
        self._sum_cache = cache_utils.load_summaries()
        self._entries.clear()

        cached_count = 0
        for f in files:
            cls = cache_utils.get_cached(f, cls_cache) or {}
            summary = cache_utils.get_summary(f, self._sum_cache)
            status = "кэш" if summary is not None else ""
            if summary is not None:
                cached_count += 1
            self._entries.append({
                "file":     f,
                "title":    (cls.get("title") or "").strip(),
                "author":   (cls.get("author") or "").strip(),
                "category": cls.get("category") or "",
                "summary":  summary or "",
                "status":   status,
            })

        need = len(files) - cached_count
        self._stats_label.configure(
            text=f"{len(files)} файлов  |  саммари: {cached_count}  |  осталось: {need}")
        self._reset_btn.configure(state="normal" if files else "disabled")
        self._export_btn.configure(state="normal" if files else "disabled")
        self._apply_tree()
        self._update_run_btn()

        if need == 0 and cached_count > 0:
            self._status_label.configure(
                text="Все саммари готовы. Нажмите «Сбросить саммари» для повторной обработки.")

    # ---------------------------------------------------------------- summarize

    def _update_run_btn(self):
        can = bool(self._entries) and any(e["status"] == "" for e in self._entries)
        self._run_btn.configure(state="normal" if can else "disabled")

    def _start_summarize(self):
        to_do = [e for e in self._entries if e["status"] == ""]
        if not to_do:
            return

        self._gen += 1
        gen = self._gen
        self._summarizer = Summarizer()
        self._run_btn.configure(state="disabled")
        self._stop_btn.configure(state="normal")
        self._progress.set(0)

        threading.Thread(
            target=self._run_summarize, args=(to_do, gen), daemon=True
        ).start()

    def _run_summarize(self, to_do: list[dict], gen: int):
        path_to_idx = {e["file"].full_path: i for i, e in enumerate(self._entries)}
        self._summarizer.summarize_all(
            files=[e["file"] for e in to_do],
            cache=self._sum_cache,
            on_result=lambda i, info, summary, from_cache:
                self._queue.put(("result", gen, path_to_idx[info.full_path], summary, from_cache)),
            on_progress=lambda cur, tot:
                self._queue.put(("progress", gen, cur, tot)),
            on_done=lambda _cache:
                self._queue.put(("done", gen)),
            on_error=lambda fp, err:
                self._queue.put(("err", gen, fp, err)),
        )

    def _stop_summarize(self):
        if self._summarizer:
            self._summarizer.stop()
        self._stop_btn.configure(state="disabled")

    def _reset_summaries(self):
        self._sum_cache = {}
        cache_utils.save_summaries(self._sum_cache)
        for e in self._entries:
            e["summary"] = ""
            e["status"] = ""
        n = len(self._entries)
        self._stats_label.configure(
            text=f"{n} файлов  |  саммари: 0  |  осталось: {n}")
        self._progress.set(0)
        self._status_label.configure(text="Саммари сброшены. Нажмите «Составить саммари».")
        self._apply_tree()
        self._update_run_btn()
        self._set_summary_text("")

    # ---------------------------------------------------------------- queue pump

    def _poll_queue(self):
        try:
            for _ in range(30):
                msg = self._queue.get_nowait()
                kind = msg[0]

                if kind == "progress":
                    _, gen, cur, tot = msg
                    if gen != self._gen:
                        continue
                    self._progress.set(cur / tot if tot else 0)
                    errors = sum(1 for e in self._entries if e["status"] == "ошибка")
                    err_str = f"  |  ошибок: {errors}" if errors else ""
                    self._status_label.configure(
                        text=f"Обрабатывается: {cur} / {tot}{err_str}")

                elif kind == "result":
                    _, gen, idx, summary, from_cache = msg
                    if gen != self._gen:
                        continue
                    e = self._entries[idx]
                    is_error = summary.startswith("Ошибка:")
                    e["summary"] = summary
                    e["status"] = "кэш" if from_cache else ("ошибка" if is_error else "ИИ")
                    self._update_row(idx)
                    done = sum(1 for e in self._entries if e["status"])
                    self._stats_label.configure(
                        text=f"{len(self._entries)} файлов  |  обработано: {done}")
                    if idx == self._selected_idx:
                        self._set_summary_text(summary)

                elif kind == "done":
                    _, gen = msg
                    if gen != self._gen:
                        continue
                    self._progress.set(1.0)
                    done  = sum(1 for e in self._entries if e["status"])
                    errors = sum(1 for e in self._entries if e["status"] == "ошибка")
                    self._status_label.configure(
                        text=f"Завершено. Обработано: {done}"
                        + (f"  |  ошибок: {errors}" if errors else ""))
                    self._stop_btn.configure(state="disabled")
                    self._export_btn.configure(state="normal")
                    self._update_run_btn()

                elif kind == "err":
                    pass

        except queue.Empty:
            pass
        finally:
            self.after(150, self._poll_queue)

    # ---------------------------------------------------------------- table

    def _apply_tree(self):
        for item in self._tree.get_children():
            self._tree.delete(item)
        self._iid_to_idx.clear()
        self._idx_to_iid.clear()
        for idx, entry in enumerate(self._entries):
            iid = self._insert_row(entry)
            self._iid_to_idx[iid] = idx
            self._idx_to_iid[idx] = iid

    def _insert_row(self, entry: dict) -> str:
        return self._tree.insert(
            "", "end",
            values=self._row_values(entry),
            tags=(self._row_tag(entry),),
        )

    def _update_row(self, idx: int):
        iid = self._idx_to_iid.get(idx)
        if iid:
            self._tree.item(iid,
                            values=self._row_values(self._entries[idx]),
                            tags=(self._row_tag(self._entries[idx]),))

    @staticmethod
    def _row_values(e: dict) -> tuple:
        preview = e["summary"].replace("\n", " ")[:100]
        return (
            e["title"] or "—",
            e["author"] or "—",
            e["category"] or "—",
            e["file"].name,
            e["file"].extension,
            preview,
            e["status"],
        )

    @staticmethod
    def _row_tag(e: dict) -> str:
        return {"кэш": "tag_cache", "ИИ": "tag_ai",
                "ошибка": "tag_error"}.get(e["status"], "")

    _sort_rev: dict[str, bool] = {}

    def _sort_by(self, col: str):
        rev = self._sort_rev.get(col, False)
        items = [(self._tree.set(k, col), k) for k in self._tree.get_children("")]
        items.sort(key=lambda x: x[0].lower(), reverse=rev)
        for i, (_, k) in enumerate(items):
            self._tree.move(k, "", i)
        self._sort_rev[col] = not rev

    # ---------------------------------------------------------------- right panel

    def _on_select(self, _event):
        sel = self._tree.selection()
        if not sel:
            return
        iid = sel[0]
        idx = self._iid_to_idx.get(iid, -1)
        if idx < 0:
            return
        self._selected_idx = idx
        e = self._entries[idx]
        self._file_label.configure(text=e["file"].full_path)
        self._set_summary_text(e["summary"])

    def _set_summary_text(self, text: str):
        self._summary_box.configure(state="normal")
        self._summary_box.delete("1.0", "end")
        self._summary_box.insert("1.0", text if text else "(саммари ещё не составлено)")
        self._summary_box.configure(state="disabled")

    # ---------------------------------------------------------------- export

    def _export_excel(self):
        if not self._entries:
            return
        path = filedialog.asksaveasfilename(
            defaultextension=".xlsx",
            filetypes=[("Excel файл", "*.xlsx")],
            initialfile="summaries.xlsx",
            title="Сохранить саммари как…",
        )
        if not path:
            return
        try:
            self._write_excel(path)
            self._status_label.configure(text=f"Экспорт завершён: {path}")
            try:
                os.startfile(path)
            except Exception:
                pass
        except Exception as exc:
            self._status_label.configure(text=f"Ошибка экспорта: {exc}")

    def _write_excel(self, path: str):
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter

        wb = Workbook()
        ws = wb.active
        ws.title = "Саммари"

        header_font  = Font(name="Calibri", bold=True, color="FFFFFF", size=11)
        header_fill  = PatternFill("solid", fgColor="1F538D")
        header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell_align   = Alignment(vertical="top", wrap_text=True)
        thin         = Side(style="thin", color="C0C0C0")
        border       = Border(left=thin, right=thin, top=thin, bottom=thin)
        alt_fill     = PatternFill("solid", fgColor="F7FBFF")
        white_fill   = PatternFill("solid", fgColor="FFFFFF")
        err_fill     = PatternFill("solid", fgColor="FDECEA")
        base_font    = Font(name="Calibri", size=10)

        headers = ["№", "Название", "Автор", "Категория", "Формат", "Имя файла", "Саммари", "Статус"]
        ws.append(headers)
        for col, _ in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col)
            cell.font      = header_font
            cell.fill      = header_fill
            cell.alignment = header_align
            cell.border    = border
        ws.row_dimensions[1].height = 30

        for row_idx, e in enumerate(self._entries, 2):
            f = e["file"]
            ws.append([
                row_idx - 1,
                e["title"] or "",
                e["author"] or "",
                e["category"] or "",
                f.extension,
                f.name,
                e["summary"] or "",
                e["status"],
            ])
            row_fill = err_fill if e["status"] == "ошибка" \
                else (white_fill if row_idx % 2 == 0 else alt_fill)
            for col in range(1, len(headers) + 1):
                cell = ws.cell(row=row_idx, column=col)
                cell.fill      = row_fill
                cell.border    = border
                cell.font      = base_font
                cell.alignment = cell_align
            ws.row_dimensions[row_idx].height = 60

        ws.freeze_panes = "A2"
        ws.auto_filter.ref = ws.dimensions

        for col, w in enumerate([5, 40, 25, 30, 8, 35, 80, 8], 1):
            ws.column_dimensions[get_column_letter(col)].width = w

        wb.save(path)
