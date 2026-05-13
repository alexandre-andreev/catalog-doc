"""Export/import taxonomy to/from Excel and persist custom taxonomy as JSON."""

import json
from pathlib import Path

CUSTOM_FILE = Path(__file__).parent.parent / "taxonomy_custom.json"


def export_taxonomy(path: str, categories: list[str]) -> None:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    wb = Workbook()
    ws = wb.active
    ws.title = "Таксономия"

    hdr_font  = Font(name="Calibri", bold=True, color="FFFFFF", size=11)
    hdr_fill  = PatternFill("solid", fgColor="1F538D")
    hdr_align = Alignment(horizontal="center", vertical="center")
    thin      = Side(style="thin", color="BBBBBB")
    border    = Border(left=thin, right=thin, top=thin, bottom=thin)
    cell_font  = Font(name="Calibri", size=10)
    cell_align = Alignment(vertical="center")
    fills = [PatternFill("solid", fgColor="EBF3FB"), PatternFill("solid", fgColor="FFFFFF")]

    ws.append(["Раздел", "Подраздел"])
    for col in range(1, 3):
        c = ws.cell(row=1, column=col)
        c.font, c.fill, c.alignment, c.border = hdr_font, hdr_fill, hdr_align, border
    ws.row_dimensions[1].height = 22

    section_fill_idx: dict[str, int] = {}
    fill_counter = 0

    for row_num, cat in enumerate(categories, 2):
        section, sub = (cat.split("/", 1) if "/" in cat else (cat, ""))

        if section not in section_fill_idx:
            section_fill_idx[section] = fill_counter % 2
            fill_counter += 1
        fill = fills[section_fill_idx[section]]

        ws.append([section, sub])
        for col in range(1, 3):
            c = ws.cell(row=row_num, column=col)
            c.font, c.fill, c.alignment, c.border = cell_font, fill, cell_align, border

    ws.column_dimensions["A"].width = 34
    ws.column_dimensions["B"].width = 34
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:B{1 + len(categories)}"

    # Instructions sheet
    ws2 = wb.create_sheet("Инструкция")
    instructions = [
        ["Как редактировать таксономию"],
        [""],
        ["Лист «Таксономия» — список категорий для AI-классификации документов."],
        ["Каждая строка — одна категория."],
        [""],
        ["Колонки:"],
        ["  A — Раздел    (обязательно): основная группа, например «Программирование»"],
        ["  B — Подраздел (необязательно): уточнение, например «Python»"],
        [""],
        ["Итоговая категория = «Раздел/Подраздел» (если подраздел задан) или «Раздел»."],
        [""],
        ["Правила:"],
        ["  • Строку 1 (заголовок) не изменяйте."],
        ["  • Добавляйте новые строки в любое место."],
        ["  • Удаляйте ненужные строки целиком."],
        ["  • Пустые строки (пустая колонка A) игнорируются."],
        [""],
        ["После редактирования: вкладка «Настройки» → «Загрузить из Excel» → выберите этот файл."],
        ["Изменения вступят в силу при следующей классификации."],
    ]
    for row in instructions:
        ws2.append(row)
    ws2.column_dimensions["A"].width = 72
    ws2.cell(row=1, column=1).font = Font(name="Calibri", bold=True, size=12)

    wb.save(path)


def load_taxonomy(path: str) -> tuple[list[str], str]:
    """Read taxonomy from Excel; return (categories_ui, category_guide)."""
    from openpyxl import load_workbook

    wb = load_workbook(path, data_only=True)
    if "Таксономия" not in wb.sheetnames:
        raise ValueError("Лист «Таксономия» не найден в файле")
    ws = wb["Таксономия"]

    categories: list[str] = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        section = str(row[0]).strip() if row[0] else ""
        sub     = str(row[1]).strip() if len(row) > 1 and row[1] else ""
        if not section:
            continue
        categories.append(f"{section}/{sub}" if sub else section)

    if not categories:
        raise ValueError("Лист «Таксономия» пуст — ни одной строки с данными")

    guide = _build_guide(categories)
    return categories, guide


def _build_guide(categories: list[str]) -> str:
    lines: list[str] = []
    current_section: str | None = None
    section_items: list[str] = []

    def flush():
        if section_items:
            lines.append(" · ".join(section_items))

    for cat in categories:
        if "/" in cat:
            section = cat.split("/", 1)[0]
            if section != current_section:
                flush()
                current_section = section
                section_items = [cat]
            else:
                section_items.append(cat)
        else:
            flush()
            current_section = None
            section_items = []
            lines.append(cat)

    flush()
    return "\n".join(lines)


def save_custom(categories: list[str], guide: str) -> None:
    CUSTOM_FILE.write_text(
        json.dumps({"categories": categories, "guide": guide},
                   ensure_ascii=False, indent=2),
        encoding="utf-8",
    )


def load_custom() -> tuple[list[str], str] | None:
    if not CUSTOM_FILE.exists():
        return None
    try:
        data = json.loads(CUSTOM_FILE.read_text(encoding="utf-8"))
        return data["categories"], data["guide"]
    except Exception:
        return None


def reset_custom() -> None:
    if CUSTOM_FILE.exists():
        CUSTOM_FILE.unlink()


def custom_exists() -> bool:
    return CUSTOM_FILE.exists()
